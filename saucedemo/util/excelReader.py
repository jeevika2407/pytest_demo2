# import openpyxl

# def get_data(path, sheet_name):
#     final_list = []
#     workbook = openpyxl.load_workbook(path)
#     sheet = workbook[sheet_name]

#     for r in range(2, sheet.max_row + 1): 
#         row_list = []
#         for c in range(1, sheet.max_column + 1):
#             row_list.append(sheet.cell(row=r, column=c).value)
#         final_list.append(tuple(row_list))  
#     return final_list


# import openpyxl

# def get_data(file, sheet):
#     workbook = openpyxl.load_workbook(file)
#     worksheet = workbook[sheet]
#     data = []

#     for i in range(2, worksheet.max_row + 1):  # Skip header
#         username = worksheet.cell(row=i, column=1).value
#         password = worksheet.cell(row=i, column=2).value
#         test_case = worksheet.cell(row=i, column=3).value
#         data.append((username, password, test_case))

#     return data



import openpyxl
def get_data(path, sheet):
    data_list=[]
    wkbk=openpyxl.load_workbook(path)
    sht=wkbk[sheet]
    tot_r=sht.max_row
    tot_c=sht.max_column
    for r in range(2, tot_r + 1):
        row_list=[]
        for c in range(1, tot_c + 1):
            row_list.append(sht.cell(row=r, column=c).value)
        data_list.append(row_list)
    return data_list

