import openpyxl

def get_data(path, sheet_name):
    final_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]

    for r in range(2, sheet.max_row + 1): 
        row_list = []
        for c in range(1, sheet.max_column + 1):
            row_list.append(sheet.cell(row=r, column=c).value)
        final_list.append(tuple(row_list))  # Convert to tuple for pytest
    return final_list
